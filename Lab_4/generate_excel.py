import csv
import openpyxl # type: ignore
from datetime import datetime

# Функція для обчислення віку
def calculate_age(birthdate):
    today = datetime.today()
    return today.year - birthdate.year - ((today.month, today.day) < (birthdate.month, birthdate.day))

def create_excel_from_csv():
    try:
        with open('employees.csv', mode='r', encoding='utf-8') as file:
            reader = csv.reader(file)
            data = list(reader)

            workbook = openpyxl.Workbook()
            sheets = {
                "all": workbook.active,
                "younger_18": workbook.create_sheet("younger_18"),
                "18-45": workbook.create_sheet("18-45"),
                "45-70": workbook.create_sheet("45-70"),
                "older_70": workbook.create_sheet("older_70")
            }
            sheets["all"].title = "all"

            # Записуємо заголовки
            headers = data[0] + ["Вік"]
            for sheet in sheets.values():
                sheet.append(headers)

            # Записуємо дані в таблиці
            for row in data[1:]:
                birthdate = datetime.strptime(row[4], '%Y-%m-%d')
                age = calculate_age(birthdate)
                new_row = row + [age]
                sheets["all"].append(new_row)

                if age < 18:
                    sheets["younger_18"].append(new_row)
                elif 18 <= age <= 45:
                    sheets["18-45"].append(new_row)
                elif 45 < age <= 70:
                    sheets["45-70"].append(new_row)
                else:
                    sheets["older_70"].append(new_row)

            workbook.save('employees.xlsx')
            print("Excel файл успішно створено.")
    except FileNotFoundError:
        print("Помилка: файл employees.csv не знайдено.")
    except Exception as e:
        print(f"Помилка при створенні Excel файлу: {e}")

if __name__ == '__main__':
    create_excel_from_csv()
